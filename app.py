import openpyxl
import os
import re
import time
import logging
from flask import Flask, request, render_template, redirect, url_for, flash, send_from_directory, session
from werkzeug.utils import secure_filename
import datetime
from analyzer import analyze
from flask_cors import CORS
from downloader import output_upload_to_gcs


tempo_totale = 0
trascrizioni_complete = 0
output_bucket_name = 'output-trascrizioni'

app = Flask(__name__)
CORS(app)
app.secret_key = os.urandom(24)
prompt_type_global = "default"

# Directory per i file caricati
UPLOAD_FOLDER = './uploads'
ALLOWED_EXTENSIONS = {'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['GCS_OUTPUT_BUCKET_NAME'] = output_bucket_name

links = []
transcriptions = []
current_file_name = None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def carica_link(file_path):
    global links
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    links.clear()

    url_regex = re.compile(r'https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+(?:/[\w\-._~!$&\'()*+,;=:@]+)*/?(?:\?[\w\-._~!$&\'()*+,;=:@/?]*)?')

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                if isinstance(cell.value, str):
                    url_match = url_regex.search(cell.value)
                    if url_match:
                        links.append(url_match.group(0))


@app.route('/')
def upload_file():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def handle_upload():
    global current_file_name

    if 'file' not in request.files:
        flash('Nessun file selezionato')
        return redirect(url_for('upload_file'))

    file = request.files['file']
    prompt_type = request.form.get('prompt_type', 'default')

    if file.filename == '' or not allowed_file(file.filename):
        flash('File non valido o formato non supportato')
        return redirect(url_for('upload_file'))

    filename = secure_filename(file.filename)
    current_file_name = filename

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)

    carica_link(file_path)
    flash('File caricato e link caricati correttamente')
    return redirect(url_for('transcriber'))

@app.route('/transcriber')
def transcriber():
    prompt_type = session.get("prompt_type")
    return render_template('index.html', links=links, transcriptions=transcriptions, prompt_type = prompt_type_global)

def save_transcriptions_to_excel(transcriptions, original_file_path, output_file_path):
    wb = openpyxl.load_workbook(original_file_path)
    sheet = wb.active

    # Trova la colonna dei link usando url_regex
    link_col = None
    url_regex = re.compile(r'https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+(?:/[\w\-._~!$&\'()*+,;=:@]+)*/?(?:\?[\w\-._~!$&\'()*+,;=:@/?]*)?')

    for col in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str) and url_regex.search(cell_value):
                link_col = col
                break
        if link_col:
            break

    if link_col is None:
        raise ValueError("Nessuna colonna trovata contenente link nel file Excel")

    # Scrivi le trascrizioni nella colonna successiva
    transcription_col = link_col + 1

    is_first_pair = True
    for link, transcription in transcriptions:
        if is_first_pair:
            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=link_col).value
                if cell_value == link:
                    sheet.cell(row=row-1, column=link_col).value = "Link"
                    sheet.cell(row=row-1, column=transcription_col).value = "Elaborazione"
                    break
                is_first_pair = False

        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=link_col).value
            if cell_value == link:
                sheet.cell(row=row, column=transcription_col).value = transcription
                break

    # Assicurati che la directory di output esista
    os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
    # Utilizza percorso assoluto
    absolute_output_file_path = os.path.abspath(output_file_path)

    # Prova a salvare il file
    try:
        wb.save(absolute_output_file_path)
    except PermissionError as e:
        print(f"Errore: {e}")
        raise PermissionError(f"Non è stato possibile salvare il file: {absolute_output_file_path}")

@app.route('/trascrizione_tutti', methods=['GET'])
def trascrizione_tutti():
    global transcriptions, tempo_totale, trascrizioni_complete
    transcriptions.clear()
    tempo_totale = 0
    trascrizioni_complete = 0
    prompt_type = prompt_type_global

    output_dir = "outputs"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    now = datetime.datetime.now()
    date_string = now.strftime("%Y%m%d")

    for link in links:
        inizio = time.time()
        max_retries = 3
        retry_delay = 10  # seconds

        if not prompt_type:
            logging.debug("prompt_type non è definito, uso il valore predefinito 'default'")
            prompt_type = 'default'
        for attempt in range(max_retries):
            try:
                elaborazione = analyze(link, prompt_type)
                transcriptions.append((link, elaborazione))
                break  # If successful, exit the retry loop
            except Exception as e:
                if attempt < max_retries - 1:  # If there are more retries left
                    logging.debug(f"Attempt {attempt + 1} failed. Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                else:
                    logging.debug(f"All {max_retries} attempts failed. Appending error message. error: {e}")
                    transcriptions.append((link, f"errore: {e}"))

        durata = time.time() - inizio
        tempo_totale += durata
        trascrizioni_complete += 1


    original_file_path = os.path.join(app.config['UPLOAD_FOLDER'], current_file_name)
    output_file_path = os.path.join(output_dir, f"analyzed_{date_string}_{current_file_name}")
    save_transcriptions_to_excel(transcriptions, original_file_path, output_file_path)

    gcs_file_name = f"analyzed_{date_string}_{current_file_name}"
    gcs_url = output_upload_to_gcs(app.config['GCS_OUTPUT_BUCKET_NAME'], output_file_path, gcs_file_name, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    logging.debug(f"Link output: {gcs_url}")

    flash('Trascrizione completata per tutti i link')
    return send_from_directory(output_dir, os.path.basename(output_file_path), as_attachment=True)

@app.route('/get_transcriptions', methods=['GET'])
def get_transcriptions():
    return {'transcriptions': transcriptions}

@app.route('/get_progress', methods=['GET'])
def get_progress():
    completate = len(transcriptions)
    totale = len(links)
    return {'completate': completate, 'totale': totale}

@app.route('/get_estimated_time', methods=['GET'])
def get_estimated_time():
    if trascrizioni_complete == 0:
        tempo_medio = 0
    else:
        tempo_medio = tempo_totale / trascrizioni_complete

    trascrizioni_rimanenti = max(len(links) - trascrizioni_complete, 0)
    tempo_rimanente = round(tempo_medio * trascrizioni_rimanenti)

    return {'tempo_rimanente': tempo_rimanente}

@app.route('/sidebar')
def show_sidebar():
    output_dir = os.path.join(app.root_path, 'outputs')
    files = os.listdir(output_dir)
    return render_template('sidebar.html', files=files)

@app.route('/download/<path:filename>')
def download_file(filename):
    output_dir = os.path.join(app.root_path, 'outputs')
    return send_from_directory(output_dir, os.path.basename(filename), as_attachment=True)


@app.route('/prompt', methods=['POST'])
def process():
    global prompt_type_global
    prompt_type_global = request.form.get('prompt_type', 'default')
    flash(f'Prompt type selected: {prompt_type_global}')
    return redirect(url_for('transcriber'))

if __name__ == '__main__':
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    app.run(host='0.0.0.0', debug=True)